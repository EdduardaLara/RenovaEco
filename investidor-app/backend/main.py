from fastapi import FastAPI, UploadFile, File
from fastapi.responses import FileResponse, JSONResponse
import pandas as pd
import os
import shutil
import uuid
from openpyxl import load_workbook
from fastapi.middleware.cors import CORSMiddleware


app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


@app.get("/")
def home():
    return {"status": "API rodando 🚀"}


def limpar_numero(valor):
    """Converte valores brasileiros (R$ 1.234,56) para float."""
    if pd.isna(valor):
        return 0.0
    return float(
        str(valor)
        .replace("R$", "")
        .replace(" ", "")
        .replace(".", "")
        .replace(",", ".")
        .strip()
    )


def encontrar_coluna(df, *palavras_chave):
    """Busca coluna que contenha TODAS as palavras-chave (case-insensitive)."""
    for col in df.columns:
        col_norm = col.lower()
        if all(p.lower() in col_norm for p in palavras_chave):
            return col
    raise ValueError(
        f"Coluna com palavras-chave {palavras_chave} não encontrada.\n"
        f"Colunas disponíveis: {list(df.columns)}"
    )


@app.post("/upload")
async def upload(file: UploadFile = File(...)):
    try:
        # 📥 Salva arquivo enviado
        file_path = os.path.join(UPLOAD_FOLDER, file.filename)
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        # 📊 Lê Excel
        df = pd.read_excel(file_path, header=0)
        df = df.dropna(how="all").reset_index(drop=True)
        df.columns = df.columns.astype(str).str.strip()

        # 🔎 Mapeia colunas
        col_compensacao = encontrar_coluna(df, "Compensa")   # G → B3
        col_saldo       = encontrar_coluna(df, "Saldo")      # H → B4
        col_boleto      = encontrar_coluna(df, "Boleto")     # L → B5
        # B6 (Inadimplentes) não existe no Excel → fica 0

        # 🔥 Pega a última linha com dados válidos
        linha = df.iloc[-1]

        kwh_compensado = limpar_numero(linha[col_compensacao])
        saldo_kwh      = limpar_numero(linha[col_saldo])
        boleto         = limpar_numero(linha[col_boleto])

        # 📄 Copia modelo base
        novo_arquivo = f"relatorio_{uuid.uuid4().hex}.xlsx"
        shutil.copy("modelo.xlsx", novo_arquivo)

        wb = load_workbook(novo_arquivo, data_only=False)
        ws = wb.active

        # ✏️ Preenche apenas as células de entrada
        # B7, B8, B9, B11 já têm valores/fórmulas fixas no modelo — não tocar!
        ws["B3"] = kwh_compensado   # kWh Compensado pelo Consumidor
        ws["B4"] = saldo_kwh        # Saldo Total do Consumidor (kWh)
        ws["B5"] = boleto           # Faturas RenovaEco Cobradas e Recebidas
        ws["B6"] = 0                # Faturas Inadimplentes — ajuste se tiver coluna no Excel

        wb.save(novo_arquivo)

        return FileResponse(
            novo_arquivo,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename="relatorio_investidor.xlsx"
        )

    except ValueError as ve:
        return JSONResponse(status_code=422, content={"erro": str(ve)})
    except Exception as e:
        return JSONResponse(status_code=500, content={"erro": str(e)})