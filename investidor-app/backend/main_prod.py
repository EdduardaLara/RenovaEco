from fastapi import FastAPI, UploadFile, File
from fastapi.responses import FileResponse, JSONResponse, HTMLResponse
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd
import os
import shutil
import uuid
from openpyxl import load_workbook

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/", response_class=HTMLResponse)
def home():
    with open("index.html", "r", encoding="utf-8") as f:
        return f.read()

def limpar_numero(valor):
    if pd.isna(valor):
        return 0.0
    s = str(valor).replace("R$", "").replace(" ", "").strip()
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    elif "." in s:
        partes = s.split(".")
        if len(partes[-1]) == 3:
            s = s.replace(".", "")
    return round(float(s), 2)

def encontrar_coluna(df, *palavras_chave):
    for col in df.columns:
        col_norm = col.lower()
        if all(p.lower() in col_norm for p in palavras_chave):
            return col
    raise ValueError(f"Coluna {palavras_chave} não encontrada. Disponíveis: {list(df.columns)}")

UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

@app.post("/upload")
async def upload(file: UploadFile = File(...)):
    try:
        file_path = os.path.join(UPLOAD_FOLDER, file.filename)
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        df = pd.read_excel(file_path, header=0)
        df = df.dropna(how="all").reset_index(drop=True)
        df.columns = df.columns.astype(str).str.strip()

        col_compensacao = encontrar_coluna(df, "Compensa")
        col_saldo       = encontrar_coluna(df, "Saldo")
        col_boleto      = encontrar_coluna(df, "Boleto")

        linha = df.iloc[-1]

        kwh_compensado = limpar_numero(linha[col_compensacao])
        saldo_kwh      = limpar_numero(linha[col_saldo])
        boleto         = limpar_numero(linha[col_boleto])

        novo_arquivo = f"relatorio_{uuid.uuid4().hex}.xlsx"
        shutil.copy("modelo.xlsx", novo_arquivo)

        wb = load_workbook(novo_arquivo, data_only=False)
        ws = wb.active

        ws["B3"] = kwh_compensado
        ws["B4"] = saldo_kwh
        ws["B5"] = boleto
        ws["B6"] = 0.0

        wb.save(novo_arquivo)

        return FileResponse(
            novo_arquivo,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename="relatorio_investidor.xlsx"
        )

    except ValueError as ve:
        return JSONResponse(status_code=422, content={"erro": str(ve)})
    except Exception as e:
        return JSONResponse(status_code=500, content={"erro": str(e)})