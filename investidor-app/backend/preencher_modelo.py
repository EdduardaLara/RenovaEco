from openpyxl import load_workbook
import shutil
import uuid

def preencher_planilha(dados_extraidos):

    # Cria cópia única do modelo
    novo_arquivo = f"relatorio_{uuid.uuid4().hex}.xlsx"
    shutil.copy("modelo.xlsx", novo_arquivo)

    wb = load_workbook(novo_arquivo)
    ws = wb.active

    # 🔥 AQUI VOCÊ DEFINE AS CÉLULAS FIXAS DO MODELO

    ws["B4"] = dados_extraidos["kwh_compensado"]
    ws["B5"] = dados_extraidos["saldo_total"]
    ws["B6"] = dados_extraidos["faturas_recebidas"]
    ws["B8"] = dados_extraidos["tarifa_base"]

    # NÃO mexe nas células:
    # B9 (economia)
    # B10 (taxa 13%)
    # B12 (repasse líquido)

    wb.save(novo_arquivo)

    return novo_arquivo
